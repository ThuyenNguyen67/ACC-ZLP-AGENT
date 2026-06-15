#!/usr/bin/env python3
"""
reconcile.py — Giai đoạn 2 (deterministic) của plugin cashbook-reconciliation.

Đọc dữ liệu ĐÃ trích xuất (input_1 = Payment, input_2 = sao kê) ở dạng JSON, đọc
File 3 (báo cáo sổ quỹ) gốc dạng .xlsx, tính cash-in / cash-out lũy kế đến hôm nay
và số dư cuối kỳ hôm nay, rồi:
  - ghi Cash In / Cash Out / End Balance mới vào BẢN SAO của File 3 (giữ template,
    Begin Balance bất biến);
  - sinh báo cáo cảnh báo cho tài khoản có số dư tính toán lệch với sao kê;
  - in JSON tóm tắt ra stdout.

KHÔNG tính tiền bằng tay — toàn bộ số học nằm ở đây để kết quả lặp lại & kiểm thử được.
Quy ước nghiệp vụ: xem references/business-rules.md và references/decisions.md.
"""

import argparse
import json
import sys
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook


# --------------------------------------------------------------------------- #
# Chuẩn hóa dữ liệu (mục 7 của đặc tả yêu cầu)
# --------------------------------------------------------------------------- #
def norm_text(s):
    """Bỏ dấu tiếng Việt + lowercase + gộp khoảng trắng."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.lower().split())


def to_amount(x):
    """Ép số tiền về int VND. Bỏ dấu phẩy ngăn nghìn, khoảng trắng, hậu tố VND.
    Ô trống / None / NaN -> 0."""
    if x is None:
        return 0
    if isinstance(x, bool):
        return 0
    if isinstance(x, (int, float)):
        # loại NaN
        if isinstance(x, float) and x != x:
            return 0
        return int(round(x))
    s = str(x).strip()
    if s == "":
        return 0
    s = s.replace(",", "").replace("VND", "").replace("vnd", "")
    s = s.replace("\u00a0", "").replace(" ", "")
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


def norm_acct(x):
    """Số tài khoản dạng chuỗi: bỏ ='...' của Excel, bỏ .0 do parse số, giữ số 0 đầu."""
    if x is None:
        return ""
    s = str(x).strip()
    if s.startswith('="') and s.endswith('"'):
        s = s[2:-1]
    s = s.strip().strip('"').strip("'").strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "")


# --------------------------------------------------------------------------- #
# Đọc dữ liệu đã trích xuất
# --------------------------------------------------------------------------- #
def load_payments(path):
    """input_1: { "payments": [ {amount, description, sending_bank_number,
    receiving_bank_number}, ... ] }"""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    items = data.get("payments", data) if isinstance(data, dict) else data
    out = []
    for it in items:
        out.append({
            "amount": to_amount(it.get("amount")),
            "desc_raw": str(it.get("description", "")),
            "desc": norm_text(it.get("description")),
            "send": norm_acct(it.get("sending_bank_number")),
            "recv": norm_acct(it.get("receiving_bank_number")),
        })
    return out


def load_statements(path):
    """input_2: { "statements": [ {bank_account_number, end_balance_statement,
    begin_balance_statement?, transactions:[{amount, description}]}, ... ] }"""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    stmts = data.get("statements", data) if isinstance(data, dict) else data
    by_acct_end = {}        # acct -> end_balance_statement
    transactions = []       # [{acct, signed amount, desc_raw, desc}]
    for st in stmts:
        acct = norm_acct(st.get("bank_account_number"))
        if st.get("end_balance_statement") is not None:
            by_acct_end[acct] = to_amount(st.get("end_balance_statement"))
        raw_transactions = st.get("transactions")
        if raw_transactions is None:
            raw_transactions = st.get("credits", [])
        for c in raw_transactions:
            transactions.append({
                "acct": acct,
                "amount": to_amount(c.get("amount")),
                "desc_raw": str(c.get("description", "")),
                "desc": norm_text(c.get("description")),
            })
    return by_acct_end, transactions


# --------------------------------------------------------------------------- #
# Unified signed ledger
# --------------------------------------------------------------------------- #

def build_ledger(payments, statement_transactions):
    ledger = []
    for p in payments:
        amount = to_amount(p.get("amount"))
        if amount <= 0:
            continue
        desc_raw = str(p.get("desc_raw", ""))
        desc = norm_text(p.get("desc", desc_raw))
        send = norm_acct(p.get("send"))
        recv = norm_acct(p.get("recv"))
        if send:
            ledger.append({
                "acct": send,
                "amount": -amount,
                "desc_raw": desc_raw,
                "desc": desc,
            })
        if recv:
            ledger.append({
                "acct": recv,
                "amount": amount,
                "desc_raw": desc_raw,
                "desc": desc,
            })

    for t in statement_transactions:
        amount = to_amount(t.get("amount"))
        if amount == 0:
            continue
        desc_raw = str(t.get("desc_raw", ""))
        ledger.append({
            "acct": norm_acct(t.get("acct")),
            "amount": amount,
            "desc_raw": desc_raw,
            "desc": norm_text(t.get("desc", desc_raw)),
        })
    return ledger


def dedup_ledger(ledger):
    kept = []
    for entry in ledger:
        match = next((item for item in kept if _same_ledger_entry(item, entry)), None)
        if match is None:
            kept.append(entry.copy())
            continue
        if len(entry["desc_raw"]) > len(match["desc_raw"]):
            match["desc_raw"] = entry["desc_raw"]
            match["desc"] = entry["desc"]
    return kept


def _same_ledger_entry(a, b):
    if a["acct"] != b["acct"] or a["amount"] != b["amount"]:
        return False
    a_desc = a.get("desc", "")
    b_desc = b.get("desc", "")
    if a_desc == b_desc:
        return True
    shorter, longer = sorted((a_desc, b_desc), key=len)
    return bool(shorter) and shorter in longer


# --------------------------------------------------------------------------- #
# Đọc / ghi File 3 (giữ template) bằng openpyxl
# --------------------------------------------------------------------------- #
HEADER_KEYS = {
    "account": "bank account number",
    "begin": "begin balance",
    "cash_in": "cash in",
    "cash_out": "cash out",
    "end": "end balance",
}


def find_header_and_columns(ws):
    """Quét tìm dòng header trong File 3, trả về (header_row_index, {key: col_index})."""
    for r in range(1, ws.max_row + 1):
        norm_cells = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                norm_cells[norm_text(val)] = c
        if HEADER_KEYS["account"] in norm_cells and HEADER_KEYS["end"] in norm_cells:
            cols = {}
            for key, label in HEADER_KEYS.items():
                if label in norm_cells:
                    cols[key] = norm_cells[label]
            missing = [k for k in HEADER_KEYS if k not in cols]
            if missing:
                raise ValueError(
                    "File 3 thiếu cột bắt buộc: "
                    + ", ".join(HEADER_KEYS[m] for m in missing)
                )
            return r, cols
    raise ValueError(
        "Không tìm thấy dòng header trong File 3 "
        "(cần có 'Bank Account number' và 'End Balance')."
    )


def load_workbook_any(path):
    suffix = Path(path).suffix.lower()
    if suffix == ".xlsx":
        return load_workbook(path)
    if suffix != ".xls":
        raise ValueError("File 3 chỉ hỗ trợ định dạng .xls, .xlsx.")

    df = pd.read_excel(path, sheet_name=0, header=None, dtype=str)
    wb = Workbook()
    ws = wb.active
    for row_idx, row in enumerate(df.itertuples(index=False, name=None), start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = "" if pd.isna(value) else str(value)
    return wb


# --------------------------------------------------------------------------- #
def reconcile(payments, stmt_end, statement_transactions, file3_path, out_xlsx, tolerance=0):
    ledger = dedup_ledger(build_ledger(payments, statement_transactions))

    # gộp theo tài khoản (chỉ để tra cứu nhanh)
    wb = load_workbook_any(file3_path)
    ws = wb.active
    header_row, cols = find_header_and_columns(ws)

    accounts_processed = 0
    alerts = []
    notes = []

    for r in range(header_row + 1, ws.max_row + 1):
        acct = norm_acct(ws.cell(row=r, column=cols["account"]).value)
        if acct == "":
            continue
        accounts_processed += 1

        begin = to_amount(ws.cell(row=r, column=cols["begin"]).value)
        cash_in_y = to_amount(ws.cell(row=r, column=cols["cash_in"]).value)
        cash_out_y = to_amount(ws.cell(row=r, column=cols["cash_out"]).value)

        # các khoản tham gia tính trong ngày (marginal)
        entries = [entry for entry in ledger if entry["acct"] == acct]
        cash_in_entries = [entry for entry in entries if entry["amount"] > 0]
        cash_out_entries = [entry for entry in entries if entry["amount"] < 0]

        cash_in_today = (
            cash_in_y
            + sum(entry["amount"] for entry in cash_in_entries)
        )
        cash_out_today = (
            cash_out_y
            + sum(-entry["amount"] for entry in cash_out_entries)
        )
        end_today = begin + cash_in_today - cash_out_today

        # ghi đè 3 cột (Begin Balance giữ nguyên)
        ws.cell(row=r, column=cols["cash_in"]).value = cash_in_today
        ws.cell(row=r, column=cols["cash_out"]).value = cash_out_today
        ws.cell(row=r, column=cols["end"]).value = end_today

        ci_items = [(entry["amount"], entry["desc_raw"]) for entry in cash_in_entries]
        co_items = [(-entry["amount"], entry["desc_raw"]) for entry in cash_out_entries]

        if acct in stmt_end:
            end_stmt = stmt_end[acct]
            if abs(end_today - end_stmt) > tolerance:
                alerts.append({
                    "account": acct,
                    "end_balance_calculated": end_today,
                    "end_balance_statement": end_stmt,
                    "difference": end_today - end_stmt,
                    "cash_in_today": cash_in_today,
                    "cash_out_today": cash_out_today,
                    "cash_in_items": ci_items,
                    "cash_out_items": co_items,
                })
        else:
            notes.append(
                f"Tài khoản {acct}: không có sao kê tương ứng -> đã cập nhật "
                f"cash-in/cash-out nhưng KHÔNG đối chiếu được số dư."
            )

    wb.save(out_xlsx)
    return accounts_processed, alerts, notes


# --------------------------------------------------------------------------- #
def fmt(n):
    return f"{n:,}".replace(",", ".")


def write_alerts_md(alerts, notes, path):
    lines = ["# Báo cáo cảnh báo đối chiếu sổ quỹ", ""]
    if not alerts:
        lines.append("✅ Không có cảnh báo: mọi tài khoản có sao kê đều khớp số dư.")
    else:
        lines.append(f"⚠️ Có **{len(alerts)}** tài khoản lệch số dư cần truy vết.")
        lines.append("")
        for a in alerts:
            lines.append(f"## Tài khoản {a['account']}")
            lines.append("")
            lines.append(f"- Số dư cuối kỳ (tính toán): **{fmt(a['end_balance_calculated'])} VND**")
            lines.append(f"- Số dư cuối kỳ (sao kê): **{fmt(a['end_balance_statement'])} VND**")
            lines.append(f"- Chênh lệch (tính − sao kê): **{fmt(a['difference'])} VND**")
            lines.append(f"- Cash-in hôm nay: {fmt(a['cash_in_today'])} VND · "
                         f"Cash-out hôm nay: {fmt(a['cash_out_today'])} VND")
            lines.append("")
            lines.append(f"### Khoản tham gia tính CASH-IN ({len(a['cash_in_items'])})")
            if a["cash_in_items"]:
                lines.append("")
                lines.append("| Amount | Description |")
                lines.append("|---|---|")
                for amt, desc in a["cash_in_items"]:
                    lines.append(f"| {fmt(amt)} | {desc} |")
            else:
                lines.append("(không có)")
            lines.append("")
            lines.append(f"### Khoản tham gia tính CASH-OUT ({len(a['cash_out_items'])})")
            if a["cash_out_items"]:
                lines.append("")
                lines.append("| Amount | Description |")
                lines.append("|---|---|")
                for amt, desc in a["cash_out_items"]:
                    lines.append(f"| {fmt(amt)} | {desc} |")
            else:
                lines.append("(không có)")
            lines.append("")
    if notes:
        lines.append("## Ghi chú")
        lines.append("")
        for n in notes:
            lines.append(f"- {n}")
        lines.append("")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description="Đối chiếu & cập nhật sổ quỹ hằng ngày.")
    ap.add_argument("--payments", required=True, help="input_1 (JSON) từ File 1 Payment")
    ap.add_argument("--statements", required=True, help="input_2 (JSON) gộp các sao kê")
    ap.add_argument("--file3", required=True, help="File 3 báo cáo sổ quỹ gốc (.xls/.xlsx)")
    ap.add_argument("--out-xlsx", required=True, help="File 3 đã cập nhật (đầu ra)")
    ap.add_argument("--out-alerts", required=True, help="Báo cáo cảnh báo (.md)")
    ap.add_argument("--tolerance", type=int, default=0,
                    help="Dung sai so số dư (VND). Mặc định 0 = khớp tuyệt đối.")
    args = ap.parse_args()

    try:
        payments = load_payments(args.payments)
        stmt_end, statement_transactions = load_statements(args.statements)
        n, alerts, notes = reconcile(
            payments,
            stmt_end,
            statement_transactions,
            args.file3,
            args.out_xlsx,
            args.tolerance,
        )
        write_alerts_md(alerts, notes, args.out_alerts)
    except Exception as e:  # báo lỗi rõ ràng cho bước trích xuất xem lại
        print(json.dumps({"error": str(e)}, ensure_ascii=False))
        sys.exit(1)

    summary = {
        "accounts_processed": n,
        "alerts": [
            {
                "account": a["account"],
                "difference": a["difference"],
                "end_balance_calculated": a["end_balance_calculated"],
                "end_balance_statement": a["end_balance_statement"],
            }
            for a in alerts
        ],
        "notes": notes,
        "out_xlsx": args.out_xlsx,
        "out_alerts": args.out_alerts,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
